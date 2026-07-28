"""
업로드 문서에서 텍스트를 뽑아내는 모듈.

지원:
  .xlsx .xlsm  - openpyxl (시트별 셀 값)
  .csv         - 표준 csv
  .pdf         - pypdf (텍스트 레이어가 있는 PDF만. 스캔본은 빈 결과)
  .hwpx        - zip + XML 파싱 (가장 안정적)
  .hwp         - OLE 구조에서 BodyText 추출 (배포판에 따라 실패 가능)
  .txt .md     - 그대로

.hwp가 실패하면 사용자에게 "한글에서 .hwpx 또는 PDF로 저장 후 다시 올려주세요"를
안내합니다. .hwpx가 PDF보다 텍스트 품질이 좋으니 우선 권장합니다.
"""
from __future__ import annotations
import csv
import io
import re
import struct
import zipfile
import zlib
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class Extracted:
    filename: str
    ok: bool
    text: str = ''
    note: str = ''
    tables: list = field(default_factory=list)

    @property
    def char_count(self) -> int:
        return len(self.text)


def _from_excel(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    lines = []
    tables = []
    for ws in wb.worksheets:
        rows = []
        lines.append(f'\n### [시트] {ws.title}')
        for row in ws.iter_rows(values_only=True):
            cells = ['' if c is None else str(c).strip() for c in row]
            if not any(cells):
                continue
            rows.append(cells)
            lines.append(' | '.join(cells))
        tables.append({'sheet': ws.title, 'rows': rows})
    wb.close()
    return Extracted(path.name, True, '\n'.join(lines), tables=tables)


def _from_csv(path):
    text_lines = []
    rows = []
    for enc in ('utf-8-sig', 'cp949', 'utf-8'):
        try:
            with open(path, newline='', encoding=enc) as f:
                for row in csv.reader(f):
                    if not any(row):
                        continue
                    rows.append(row)
                    text_lines.append(' | '.join(row))
            break
        except UnicodeDecodeError:
            rows = []
            text_lines = []
            continue
    return Extracted(path.name, bool(rows), '\n'.join(text_lines), tables=[{'sheet': path.stem, 'rows': rows}])


def _from_pdf(path):
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    chunks = []
    for i, page in enumerate(reader.pages, 1):
        t = (page.extract_text() or '').strip()
        if not t:
            continue
        chunks.append(f'\n### [{i}쪽]\n{t}')
    text = '\n'.join(chunks)
    if not text.strip():
        return Extracted(path.name, False, note='텍스트 레이어가 없는 PDF입니다(스캔본으로 보입니다). 원본 한글파일을 .hwpx로 저장해서 올리거나, OCR을 거친 PDF를 올려주세요.')
    return Extracted(path.name, True, text)


def _from_hwpx(path):
    try:
        with zipfile.ZipFile(path) as z:
            names = [n for n in z.namelist() if re.match('Contents/section\\d+\\.xml$', n)]
            names.sort()
            if not names:
                return Extracted(path.name, False, note='hwpx 내부에서 본문(section) XML을 찾지 못했습니다.')
            parts = []
            for n in names:
                xml = z.read(n).decode('utf-8', errors='ignore')
                runs = re.findall('<hp:t>(.*?)</hp:t>', xml, flags=re.S)
                if not runs:
                    runs = re.findall('<[a-zA-Z]*:?t>(.*?)</[a-zA-Z]*:?t>', xml, flags=re.S)
                parts.extend(_unescape(r) for r in runs)
        text = '\n'.join(p for p in parts if p.strip())
        return Extracted(path.name, bool(text.strip()), text)
    except Exception as e:
        return Extracted(path.name, False, note=f'hwpx 해석 실패: {e}')


def _unescape(s):
    return s.replace('&lt;', '<').replace('&gt;', '>').replace('&amp;', '&').replace('&quot;', '"').replace('&#13;', '\n')


_HWPTAG_PARA_TEXT = 67


def _from_hwp(path):
    """
    HWP 5.0 OLE 구조에서 본문 텍스트를 뽑습니다.
    암호가 걸렸거나 배포용(DRM) 문서는 실패합니다. 실패 시 .hwpx 변환을 안내합니다.
    """
    import olefile
    fallback = "이 .hwp 파일에서 본문을 읽지 못했습니다. 한글에서 [다른 이름으로 저장] → 파일 형식 'HWPX'를 선택해 저장한 뒤 다시 올려주세요. (PDF도 가능하지만 hwpx 쪽이 텍스트 품질이 좋습니다.)"
    try:
        ole = olefile.OleFileIO(str(path))
    except Exception:
        return Extracted(path.name, False, note=fallback)
    try:
        header = ole.openstream('FileHeader').read()
        props = struct.unpack('<I', header[36:40])[0]
        compressed = bool(props & 1)
        encrypted = bool(props & 2)
        if encrypted:
            return Extracted(path.name, False, note='암호가 설정된 hwp 파일입니다. 암호를 푼 뒤 다시 올려주세요.')
        sections = sorted(
            (e for e in ole.listdir() if e[0] == 'BodyText'),
            key=lambda e: int(re.sub('\\D', '', e[1]) or 0),
        )
        parts = []
        for entry in sections:
            data = ole.openstream(entry).read()
            if compressed:
                try:
                    data = zlib.decompress(data, -15)
                except zlib.error:
                    continue
            parts.append(_parse_hwp_records(data))
        text = '\n'.join(p for p in parts if p.strip())
        return Extracted(path.name, bool(text.strip()), text, note='' if text.strip() else fallback)
    except Exception:
        return Extracted(path.name, False, note=fallback)
    finally:
        ole.close()


def _parse_hwp_records(data):
    """HWP 레코드 스트림을 훑어 문단 텍스트(tag 67)만 모읍니다."""
    out = []
    pos = 0
    size = len(data)
    while pos + 4 <= size:
        (hdr,) = struct.unpack('<I', data[pos:pos + 4])
        tag = hdr & 1023
        length = hdr >> 20 & 4095
        pos += 4
        if length == 4095:
            if pos + 4 > size:
                break
            (length,) = struct.unpack('<I', data[pos:pos + 4])
            pos += 4
        chunk = data[pos:pos + length]
        pos += length
        if tag == _HWPTAG_PARA_TEXT:
            out.append(_decode_para(chunk))
        if pos + 4 > size:
            break
    return '\n'.join(t for t in out if t.strip())


def _decode_para(chunk):
    """UTF-16LE 문단. 제어문자(0~31)는 인라인 오브젝트라 건너뜁니다."""
    chars = []
    i = 0
    while i + 1 < len(chunk):
        code = chunk[i] | chunk[i + 1] << 8
        if code < 32:
            i += 16 if code in (1, 2, 3, 11, 12, 14, 15, 16, 17, 18, 21, 22, 23) else 2
            if code in (10, 13):
                chars.append('\n')
            continue
        chars.append(chr(code))
        i += 2
    return ''.join(chars)


_HANDLERS = {
    '.xlsx': _from_excel,
    '.xlsm': _from_excel,
    '.csv': _from_csv,
    '.pdf': _from_pdf,
    '.hwpx': _from_hwpx,
    '.hwp': _from_hwp,
}
SUPPORTED = sorted(_HANDLERS) + ['.txt', '.md']


def extract(path):
    ext = path.suffix.lower()
    if ext in ('.txt', '.md'):
        for enc in ('utf-8', 'cp949'):
            try:
                return Extracted(path.name, True, path.read_text(encoding=enc))
            except UnicodeDecodeError:
                continue
        return Extracted(path.name, False, note='텍스트 인코딩을 판별하지 못했습니다.')
    handler = _HANDLERS.get(ext)
    if not handler:
        return Extracted(path.name, False, note=f'{ext} 형식은 아직 지원하지 않습니다. 지원 형식: {", ".join(SUPPORTED)}')
    try:
        return handler(path)
    except Exception as e:
        return Extracted(path.name, False, note=f'읽기 실패: {e}')
