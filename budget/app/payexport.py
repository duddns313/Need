"""간편결제 영수증 파일 읽기 — 뱅샐이 못 가져오는 '무엇을 샀는가'.

카드 명세서에는 네이버페이 결제가 가맹점 '네이버페이'로만 찍힌다. 무엇을
샀는지는 네이버페이 쪽에만 있다. 다행히 네이버페이는 카드영수증을 엑셀로
내려받게 해 준다. 거기엔 상호보다 나은 것이 들어 있다 — **상품명**이다.

    승인번호 | 카드사 | 카드번호 | 거래종류/할부 | 결제일자 | 취소일자
    | 상품명 | 승인금액 | 취소금액 | 공급가액 | 부가세액 | 봉사료
    | 컵보증금 | 합계

이 파일을 읽어 뱅샐 거래와 금액·날짜로 짝을 맞추고, 이름을 상품명으로
바꿔 놓는다. 짝이 없으면 그대로 알려준다 — 조용히 버리면 다 된 줄 안다.

카카오페이 거래내역서도 열 이름만 다르지 구조가 같아서 같이 읽는다.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl

# 열 이름이 앱마다 조금씩 다르다. 뜻이 같은 것끼리 묶어 둔다.
COLUMNS = {
    'date': ['결제일자', '발행일자', '거래일시', '거래일자', '이용일자', '승인일자',
             '일시', '날짜'],
    'name': ['상품명', '가맹점', '가맹점명', '내용', '거래내용', '적요', '이용처'],
    'amount': ['합계', '승인금액', '거래금액', '이용금액', '결제금액', '금액'],
    'cancel': ['취소금액'],
    'cancel_date': ['취소일자'],
    'issuer': ['카드사', '결제수단', '결제수단명'],
}

NAME_MAX = 40          # 상품명이 한 줄을 넘기면 화면에서 읽을 수가 없다
HEADER_SCAN = 12       # 머리글이 몇 줄 아래에 있을 수도 있다


class PayParseError(Exception):
    pass


@dataclass
class Receipt:
    when: datetime
    name: str
    amount: int
    issuer: str = ''
    cancelled: bool = False
    raw_name: str = ''
    sheet: str = ''

    @property
    def day(self) -> date:
        return self.when.date()


@dataclass
class PayFile:
    path: str
    rows: list[Receipt] = field(default_factory=list)
    source: str = ''
    sheets: list[str] = field(default_factory=list)
    skipped: list[str] = field(default_factory=list)

    @property
    def total(self) -> int:
        return sum(r.amount for r in self.rows if not r.cancelled)


def _find_header(ws) -> tuple[int, dict]:
    """머리글 줄과 '뜻 -> 열번호' 표를 찾는다."""
    for row in range(1, min(HEADER_SCAN, ws.max_row) + 1):
        labels = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and v.strip():
                labels[v.strip()] = col
        if not labels:
            continue
        got = {}
        for key, names in COLUMNS.items():
            for n in names:
                if n in labels:
                    got[key] = labels[n]
                    break
        if 'date' in got and 'amount' in got and 'name' in got:
            return row, got
    raise PayParseError(
        '결제일자·상품명·금액 열을 못 찾았습니다. '
        '네이버페이 카드영수증이나 카카오페이 거래내역서 파일이 맞나요?')


def _as_dt(value):
    """날짜 칸이 시트마다 다르다.

    카드영수증은 '2025-09-20 23:08:59', 현금영수증은 '20250709192100' 처럼
    붙여 쓴 열네 자리다. 둘 다 받는다.
    """
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if not value:
        return None
    text = str(value).strip()

    digits = ''.join(ch for ch in text if ch.isdigit())
    if len(digits) in (8, 12, 14) and text.replace(' ', '').isdigit():
        fmt = {8: '%Y%m%d', 12: '%Y%m%d%H%M', 14: '%Y%m%d%H%M%S'}[len(digits)]
        try:
            return datetime.strptime(digits, fmt)
        except ValueError:
            pass

    text = text.replace('.', '-').replace('/', '-')
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y%m%d'):
        try:
            return datetime.strptime(text[:len(fmt) + 4].strip(), fmt)
        except ValueError:
            continue
    return None


def _as_int(value) -> int:
    if value is None or value == '':
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).replace(',', '').replace('원', '').strip()
    try:
        return int(round(float(text)))
    except ValueError:
        return 0


def _shorten(name: str) -> str:
    """'핑기 실리카겔 재사용 제습제 습기제거제 곰팡이 방지 가정 150g 1개+가정 450g 1개'
    같은 상품명은 그대로 두면 목록이 무너진다. 앞부분만 남긴다."""
    name = ' '.join(str(name).split())
    if len(name) <= NAME_MAX:
        return name
    cut = name[:NAME_MAX]
    space = cut.rfind(' ')
    if space > NAME_MAX * 0.6:      # 낱말 중간에서 끊지 않는다
        cut = cut[:space]
    return cut + '…'


def parse(path) -> PayFile:
    """시트를 전부 읽는다.

    네이버페이는 달마다 시트를 따로 만들고, 카드영수증과 현금영수증을
    한 파일에 섞어 넣는다(열 구성도 다르다). 첫 시트만 읽으면 1년치 중
    한 달만 들어온다 — 실제로 그렇게 놓칠 뻔했다.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = PayFile(path=str(path))
    skipped = []

    for ws in wb.worksheets:
        try:
            header, cols = _find_header(ws)
        except PayParseError:
            skipped.append(ws.title)
            continue
        out.sheets.append(ws.title)
        for row in ws.iter_rows(min_row=header + 1, values_only=True):
            get = lambda key: (row[cols[key] - 1] if key in cols and cols[key] - 1 < len(row)
                               else None)
            when = _as_dt(get('date'))
            amount = _as_int(get('amount'))
            raw = get('name')
            if not when or not amount or not raw:
                continue
            out.rows.append(Receipt(
                when=when, name=_shorten(raw), raw_name=' '.join(str(raw).split()),
                amount=amount, issuer=str(get('issuer') or '').strip(),
                sheet=ws.title,
                cancelled=bool(_as_int(get('cancel'))) or bool(_as_dt(get('cancel_date'))),
            ))
    wb.close()
    out.skipped = skipped
    out.source = f'{len(out.sheets)}개 시트'
    if not out.rows:
        raise PayParseError('읽을 수 있는 거래가 한 줄도 없습니다.')
    return out


DAY_TOLERANCE = 3       # 카드 승인일과 페이 결제일이 하루이틀 어긋난다
CASH_TOLERANCE = 20     # 현금영수증은 물건 받은 뒤에 발행되기도 한다


def apply_names(transactions, receipts, only_vague=True) -> dict:
    """영수증의 상품명을 뱅샐 거래에 붙인다.

    금액이 같고 날짜가 며칠 안쪽인 것을 짝으로 본다. 한 거래에 두 번 붙지
    않게 쓴 것은 표시해 둔다. 가까운 날짜부터 가져간다.
    """
    used = set()
    filled, noted, missed = [], [], []
    lo = min((t.date for t in transactions), default=None)
    hi = max((t.date for t in transactions), default=None)
    outside = []

    def find(r, pool):
        limit = CASH_TOLERANCE if '현금' in r.sheet else DAY_TOLERANCE
        best, gap = None, 999
        for tx in pool:
            if id(tx) in used or tx.amount != r.amount:
                continue
            d = abs((tx.date - r.day).days)
            if d <= limit and d < gap:
                best, gap = tx, d
        return best

    vague = [t for t in transactions if _is_vague(t)]
    named = [t for t in transactions if not _is_vague(t)]

    for r in sorted(receipts, key=lambda x: x.when):
        if r.cancelled:
            continue
        if lo and (r.day < lo or r.day > hi):
            outside.append(r)
            continue

        # 이름이 뭉뚱그려진 거래부터 채운다. 그게 이 파일이 필요한 이유다.
        hit = find(r, vague)
        if hit is not None:
            used.add(id(hit))
            # 원본 이름을 남겨 둔다. 키가 이 값으로 만들어지므로 이름을 바꿔도
            # 사용자가 화면에서 해 둔 결정이 그대로 붙어 있는다.
            hit.orig_content = hit.orig_content or hit.content
            hit.content = r.name
            hit.memo = (hit.memo + ' ' if hit.memo else '') + r.raw_name
            filled.append((hit, r))
            continue

        # 이미 상호가 제대로 찍힌 거래라면 이름은 그대로 두고, 무엇을 샀는지만
        # 메모에 붙인다. 멀쩡한 상호를 상품명으로 덮으면 오히려 알아보기 어렵다.
        if not only_vague or True:
            hit = find(r, named)
            if hit is not None:
                used.add(id(hit))
                hit.memo = (hit.memo + ' ' if hit.memo else '') + r.raw_name
                noted.append((hit, r))
                continue

        missed.append(r)

    return {'filled': filled, 'noted': noted, 'missed': missed,
            'outside': outside, 'read': len(receipts)}


VAGUE_NAMES = {'네이버페이', '카카오페이', '토스', '페이코', '삼성페이', '애플페이',
               '스마일캐시', '네이버파이낸셜', '네이버파이낸셜㈜', '카카오페이㈜'}


def _is_vague(tx) -> bool:
    """이름만 남고 무엇을 샀는지는 안 남은 거래.

    '네이버파이낸셜'과 '네이버파이낸셜(주)'는 같은 것이다. 법인격 표기를
    떼고 본다 — 안 그러면 5건이 조용히 빠진다.
    """
    name = ' '.join(str(tx.content or '').split())
    for junk in ('(주)', '㈜', '주식회사', '(유)', '유한회사'):
        name = name.replace(junk, '')
    return name.strip() in VAGUE_NAMES
