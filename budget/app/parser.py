"""뱅크샐러드 내보내기 엑셀 파서.

뱅샐 파일은 시트 두 장으로 온다.
  - '가계부 내역' : 거래 1행씩 (날짜/시간/타입/대분류/소분류/내용/금액/화폐/결제수단/메모)
  - '뱅샐현황'    : 자산·부채·투자·대출 스냅샷

여기서는 '읽어서 그대로 담기'만 한다. 분류·상계·집계는 각각 다른 모듈의 일이다.
"""
from __future__ import annotations

import datetime as dt
import hashlib
from dataclasses import dataclass, field

import openpyxl

TX_SHEET = '가계부 내역'
STATUS_SHEET = '뱅샐현황'

EXPECTED_HEADER = ['날짜', '시간', '타입', '대분류', '소분류', '내용', '금액', '화폐', '결제수단', '메모']


class ParseError(Exception):
    """뱅샐 파일이 아니거나 형식이 바뀐 경우."""


@dataclass
class Transaction:
    date: dt.date
    time: str
    bs_type: str          # 뱅샐 타입: 수입/지출/이체
    bs_major: str         # 뱅샐 대분류
    bs_minor: str         # 뱅샐 소분류
    content: str
    amount: int           # 항상 양수. 방향은 bs_type / inflow 가 정한다
    method: str
    memo: str
    owner: str            # 남편/아내 (업로드 슬롯에서 부여)
    seq: int = 0          # 같은 파일 안에서 완전히 동일한 행의 등장 순번 (0,1,2…)
    # 원본 금액의 부호. '이체' 줄은 타입만 봐서는 들어온 건지 나간 건지 알 수가
    # 없고, 부호가 그걸 알려주는 유일한 단서다. abs() 로 지워 버리면 남에게
    # 부친 돈과 남이 보내 준 돈이 한 덩어리가 된다 — 실제로 두 파일 모두
    # '이체'가 절반씩 양수·음수로 섞여 있었다.
    inflow: bool = False

    # 분류기가 채우는 값
    category: str = ''
    nature: str = ''
    rule: str = ''        # 어떤 규칙으로 분류됐는지 (되돌리기·설명용)
    shared: bool = False  # 공동비용 여부
    offset_with: str = '' # 부부간 이체로 상계된 상대 거래 uid
    spike: str = ''       # '' 미확인 | 'once' 일회성(평소 계산 제외) | 'normal' 평소
    orig_content: str = '' # 이름을 바꾸기 전의 원본 (uid를 붙들어 두는 용도)

    @property
    def uid(self) -> str:
        """중복 업로드 제거용 키. 같은 거래는 항상 같은 값이 나온다.

        두 가지를 구분해야 한다. 둘 다 '모든 칸이 같은 두 행'으로 보이지만 처리가 반대다.

          (1) 같은 파일을 두 번 업로드  → 하나로 합쳐야 한다
          (2) 한 파일 안에 똑같은 행 두 개 → 둘 다 남겨야 한다 (진짜 거래 두 건일 수 있다)

        그래서 키에 두 가지를 넣는다.
          - `method`(결제수단·계좌): 빼면 계좌 간 이체의 양쪽 다리가 뭉개진다.
            뱅샐은 보낸 통장과 받은 통장을 각각 한 행씩 기록하는데, 둘은 날짜·시각·
            금액·내용이 같고 통장만 다르다. 실제로 이 버그로 48건이 조용히 사라졌다.
          - `seq`(파일 내 등장 순번): (2)를 살린다. 같은 파일을 두 번 넣으면 두 번째
            파일의 순번도 0,1,2… 로 같게 매겨지므로 (1)은 여전히 합쳐진다.
        """
        # 이름은 나중에 바뀔 수 있다. 네이버페이 영수증을 붙이면 '네이버페이'가
        # '핑기 실리카겔 제습제'가 된다. 그때 키까지 바뀌면 화면에 저장해 둔
        # 사용자의 결정(분류 수정·일회성 표시)이 통째로 떨어져 나간다.
        # 그래서 키는 언제나 원본 이름으로 만든다.
        raw = '|'.join([
            self.owner, str(self.date), self.time, self.bs_type,
            str(self.amount), self.orig_content or self.content,
            self.method, str(self.seq),
        ])
        return hashlib.sha1(raw.encode('utf-8')).hexdigest()[:16]

    @property
    def month(self) -> str:
        return f'{self.date.year:04d}-{self.date.month:02d}'


@dataclass
class AssetLine:
    section: str   # 자산 / 부채 / 투자 / 대출
    group: str
    name: str
    amount: int = 0
    extra: dict = field(default_factory=dict)


@dataclass
class ParsedFile:
    owner: str
    transactions: list[Transaction] = field(default_factory=list)
    assets: list[AssetLine] = field(default_factory=list)
    liabilities: list[AssetLine] = field(default_factory=list)
    investments: list[AssetLine] = field(default_factory=list)
    loans: list[AssetLine] = field(default_factory=list)

    @property
    def total_assets(self) -> int:
        return sum(a.amount for a in self.assets)

    @property
    def total_liabilities(self) -> int:
        return sum(a.amount for a in self.liabilities)

    @property
    def net_worth(self) -> int:
        return self.total_assets - self.total_liabilities


def _as_int(value) -> int:
    if value is None or value == '':
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def _as_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d'):
            try:
                return dt.datetime.strptime(value.strip()[:10], fmt).date()
            except ValueError:
                continue
    return None


def _as_time(value) -> str:
    if isinstance(value, dt.time):
        return value.strftime('%H:%M:%S')
    if isinstance(value, dt.datetime):
        return value.strftime('%H:%M:%S')
    return str(value or '')


def _text(value) -> str:
    return str(value).strip() if value is not None else ''


def parse(path, owner: str) -> ParsedFile:
    """뱅샐 엑셀 한 개를 읽는다. owner는 '남편'/'아내' 같은 소유자 라벨."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if TX_SHEET not in wb.sheetnames:
        raise ParseError(
            f"'{TX_SHEET}' 시트가 없습니다. 뱅크샐러드에서 내보낸 엑셀이 맞는지 확인해 주세요."
        )

    result = ParsedFile(owner=owner)
    result.transactions = _parse_transactions(wb[TX_SHEET], owner)
    if STATUS_SHEET in wb.sheetnames:
        _parse_status(wb[STATUS_SHEET], result)
    wb.close()
    return result


def _parse_transactions(ws, owner: str) -> list[Transaction]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return []
    header = [_text(h) for h in header[:len(EXPECTED_HEADER)]]
    if header[:3] != EXPECTED_HEADER[:3]:
        raise ParseError(
            f'거래 시트 머리글이 예상과 다릅니다: {header}. 뱅크샐러드 형식이 바뀌었을 수 있습니다.'
        )

    out = []
    seen = {}   # 같은 파일 안에서 완전히 동일한 행이 몇 번째로 나왔는지
    for row in rows:
        date = _as_date(row[0])
        if date is None:
            continue
        raw = _as_int(row[6])
        amount = abs(raw)
        if amount == 0 and not _text(row[5]):
            continue
        key = (date, _as_time(row[1]), _text(row[2]), amount, _text(row[5]),
               _text(row[8]) if len(row) > 8 else '')
        seq = seen.get(key, 0)
        seen[key] = seq + 1
        out.append(Transaction(
            date=date,
            time=_as_time(row[1]),
            bs_type=_text(row[2]),
            bs_major=_text(row[3]),
            bs_minor=_text(row[4]),
            content=_text(row[5]),
            amount=amount,
            method=_text(row[8]) if len(row) > 8 else '',
            memo=_text(row[9]) if len(row) > 9 else '',
            owner=owner,
            seq=seq,
            inflow=(raw > 0),
        ))
    return out


def _parse_status(ws, result: ParsedFile) -> None:
    """'뱅샐현황' 시트에서 자산/부채/투자/대출을 긁는다.

    이 시트는 섹션 제목(`3.재무현황` 등)으로 구간이 나뉘고 병합셀이 많아서
    행 위치를 고정으로 못 잡는다. 섹션 제목을 만나면 모드를 바꾸는 방식으로 읽는다.
    """
    mode = None
    group = ''
    for row in ws.iter_rows(values_only=True):
        cells = [_text(c) for c in row]
        joined = ' '.join(c for c in cells if c)
        if not joined:
            continue

        if '재무현황' in joined:
            mode, group = 'balance', ''
            continue
        if '투자현황' in joined:
            mode, group = 'investment', ''
            continue
        if '대출현황' in joined:
            mode, group = 'loan', ''
            continue
        if '보험현황' in joined or '고객정보' in joined or '현금흐름현황' in joined:
            mode = None
            continue
        if joined.startswith('총자산') or joined.startswith('총계') or joined.startswith('순자산'):
            continue

        values = [c for c in cells if c]
        if mode == 'balance':
            # 자산은 왼쪽 3열(항목/상품명/금액), 부채는 오른쪽 3열
            _read_balance_row(cells, result)
        elif mode == 'investment':
            item = _read_wide_row(cells, 'investment')
            if item:
                result.investments.append(item)
        elif mode == 'loan':
            item = _read_wide_row(cells, 'loan')
            if item:
                result.loans.append(item)
        elif mode is None and len(values) == 0:
            group = ''


def _read_balance_row(cells: list[str], result: ParsedFile) -> None:
    # 헤더 행은 건너뛴다
    if '상품명' in cells or '항목' in cells:
        return
    left = [c for c in cells[:5] if c]
    right = [c for c in cells[5:] if c]

    def push(chunk, bucket, section):
        if len(chunk) < 2:
            return
        amount = _as_int(chunk[-1].replace(',', ''))
        if amount == 0 and len(chunk) < 3:
            return
        name = chunk[-2]
        grp = chunk[0] if len(chunk) >= 3 else ''
        bucket.append(AssetLine(section=section, group=grp, name=name, amount=amount))

    push(left, result.assets, '자산')
    push(right, result.liabilities, '부채')


def _read_wide_row(cells: list[str], kind: str):
    values = [c for c in cells if c]
    if len(values) < 4:
        return None
    if values[0] in ('투자상품종류', '대출종류', '총계'):
        return None
    try:
        if kind == 'investment':
            principal = _as_int(values[3].replace(',', ''))
            valuation = _as_int(values[4].replace(',', '')) if len(values) > 4 else principal
            return AssetLine(
                section='투자', group=values[0], name=values[2],
                amount=valuation, extra={'원금': principal, '평가금액': valuation},
            )
        principal = _as_int(values[3].replace(',', ''))
        balance = _as_int(values[4].replace(',', '')) if len(values) > 4 else principal
        return AssetLine(
            section='대출', group=values[0], name=values[2],
            amount=balance, extra={'원금': principal, '잔액': balance},
        )
    except (IndexError, ValueError):
        return None
