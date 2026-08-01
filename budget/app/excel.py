"""올인원 관리표(.xlsx) 생성.

핵심 원칙: **값이 아니라 수식을 쓴다.**
파이썬으로 계산한 숫자를 박아넣으면, 사용자가 거래 한 줄을 손으로 고쳤을 때
숫자가 안 움직인다. 그건 가계부가 아니라 스크린샷이다.

성능 상한 (설계서 기준):
  - 수식 선반영 3,000행, 전체 열 참조(A:A) 0개
  - 휘발성 함수 0개 (TODAY도 안 쓴다)
  - Excel 2007 세대 함수만: SUMIFS COUNTIFS IFERROR EOMONTH TEXT
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import aggregate, categories as cat

ROWS = 3000                   # 수식을 미리 깔아둘 행 수
TX = '거래내역'
INFO = '기준정보'
MONTHLY = '월집계'
DASH = '대시보드'
ASSETS = '자산부채'

NAVY = 'FF1F2B50'
NAVY_LIGHT = 'FF3C4A72'
PINK = 'FFE0447D'
INK = 'FF0B0B0B'

MONEY = '₩#,##0'
PCT = '0%'
DATE = 'yyyy-mm-dd'

_thin = Side(style='thin', color='FFD9D9D9')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header(ws, row: int, labels, fill=NAVY):
    for col, text in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color='FFFFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def _input_cell(cell):
    """사용자가 채우는 칸은 분홍 테두리 + 연한 배경 (시골쥐 표 방식)."""
    cell.fill = PatternFill('solid', fgColor='FFFDF0F5')
    cell.font = Font(color='FF0B0B0B', bold=True)
    cell.border = Border(left=Side(style='thin', color=PINK),
                         right=Side(style='thin', color=PINK),
                         top=Side(style='thin', color=PINK),
                         bottom=Side(style='thin', color=PINK))


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------- 시트별
def _sheet_usage(wb, settings):
    ws = wb.create_sheet('사용법')
    _widths(ws, [4, 92])
    lines = [
        ('', ''),
        ('', '부부 공동 가계부 — 올인원 관리표'),
        ('', ''),
        ('', '■ 어디를 만지면 되나'),
        ('', '   분홍색 칸만 채우면 됩니다. 나머지는 전부 자동으로 계산됩니다.'),
        ('', f'   · {INFO} 시트  : 기준월, 시작잔액, 카테고리별 예산'),
        ('', f'   · {TX} 시트  : 거래를 손으로 더 넣고 싶을 때'),
        ('', ''),
        ('', '■ 건드리면 안 되는 곳'),
        ('', f'   {MONTHLY}, {DASH} 시트는 수식으로만 되어 있습니다. 값을 덮어쓰면 숫자가 깨집니다.'),
        ('', ''),
        ('', '■ 기준월 바꾸기'),
        ('', f'   {INFO} 시트의 기준월 칸(분홍색) 하나만 바꾸면 {DASH}가 전부 그 달 기준으로 바뀝니다.'),
        ('', ''),
        ('', '■ 거래를 더 넣으려면'),
        ('', f'   {TX} 시트 맨 아래에 이어서 쓰면 됩니다. 3,000행까지 자동으로 잡힙니다.'),
        ('', '   금액은 항상 양수로 넣고, 수입/지출 구분은 "구분" 열이 합니다.'),
        ('', '   환불은 금액을 음수로 한 줄 넣으면 해당 카테고리 합계가 그만큼 줄어듭니다.'),
        ('', ''),
        ('', '■ 이체·카드대금이 지출에 없는 이유'),
        ('', '   계좌 간 이체, 카드값 납부, 페이 충전은 쓴 돈이 아니라 돈이 옮겨간 것입니다.'),
        ('', '   구분이 "제외"로 되어 있어 지출 합계에 잡히지 않습니다. 이게 맞습니다.'),
    ]
    for row, (a, b) in enumerate(lines, start=1):
        ws.cell(row=row, column=2, value=b)
    ws['B2'].font = Font(bold=True, size=15, color=NAVY)
    for row in (4, 9, 12, 15, 20):
        ws.cell(row=row, column=2).font = Font(bold=True, size=11, color=NAVY)
    ws.sheet_view.showGridLines = False
    return ws


def _sheet_info(wb, transactions, settings):
    ws = wb.create_sheet(INFO)
    _widths(ws, [18, 20, 16, 14, 30])

    ws['A1'] = '기준정보'
    ws['A1'].font = Font(bold=True, size=14, color=NAVY)

    months = aggregate.months_of(transactions)
    latest = months[-1] if months else '2026-01'
    year, month = int(latest[:4]), int(latest[5:])

    ws['A3'] = '기준월'
    ws['B3'] = f'{year}-{month:02d}-01'
    ws['B3'].number_format = DATE
    _input_cell(ws['B3'])
    ws['C3'] = '← 이 칸만 바꾸면 대시보드가 따라 움직입니다'
    ws['C3'].font = Font(size=9, color='FF898781')

    ws['A4'] = '시작 잔액'
    ws['B4'] = int(settings.get('start_balance', 0) or 0)
    ws['B4'].number_format = MONEY
    _input_cell(ws['B4'])

    ws['A5'] = '분담 기준'
    ws['B5'] = '반반' if settings.get('split') == 'half' else '소득 비례'

    # 카테고리 + 예산
    _header(ws, 7, ['카테고리', '성격', '월 예산'])
    budgets = settings.get('budgets') or {}
    row = 8
    for category in cat.ORDER:
        if category == '미분류':
            continue
        ws.cell(row=row, column=1, value=category).border = BORDER
        ws.cell(row=row, column=2, value=cat.nature_of(category)).border = BORDER
        budget_cell = ws.cell(row=row, column=3, value=int(budgets.get(category, 0)))
        budget_cell.number_format = MONEY
        _input_cell(budget_cell)
        row += 1

    ws['E7'] = '소유자'
    ws['E7'].font = Font(bold=True, color=NAVY)
    ws['E8'] = settings.get('husband_name', '남편')
    ws['E9'] = settings.get('wife_name', '아내')

    ws['E11'] = '구분(성격)'
    ws['E11'].font = Font(bold=True, color=NAVY)
    for i, nature in enumerate(cat.NATURES):
        ws.cell(row=12 + i, column=5, value=nature)

    return ws, row - 1


def _sheet_transactions(wb, transactions, settings):
    ws = wb.create_sheet(TX)
    _widths(ws, [12, 10, 10, 14, 34, 14, 20, 9, 12])
    _header(ws, 1, ['날짜', '소유자', '구분', '카테고리', '내용', '금액',
                    '결제수단', '공동/개인', '월키'])
    ws.freeze_panes = 'A2'

    for i, tx in enumerate(sorted(transactions, key=lambda t: (t.date, t.time))):
        row = i + 2
        if row > ROWS + 1:
            break
        ws.cell(row=row, column=1, value=tx.date).number_format = DATE
        ws.cell(row=row, column=2, value=tx.owner)
        ws.cell(row=row, column=3, value=tx.nature)
        ws.cell(row=row, column=4, value=tx.category)
        ws.cell(row=row, column=5, value=tx.content)
        ws.cell(row=row, column=6, value=tx.amount).number_format = MONEY
        ws.cell(row=row, column=7, value=tx.method)
        ws.cell(row=row, column=8, value='공동' if tx.shared else '개인')

    # 월키는 마지막 행까지 수식으로 깔아둔다 — 손으로 행을 추가해도 잡히게
    for row in range(2, ROWS + 2):
        cell = ws.cell(row=row, column=9)
        cell.value = f'=IF($A{row}="","",EOMONTH($A{row},0))'
        cell.number_format = DATE

    # 드롭다운 (목록은 기준정보 시트를 참조)
    last_category_row = 7 + len([c for c in cat.ORDER if c != '미분류'])
    validations = [
        (f'$D$2:$D${ROWS + 1}', f"'{INFO}'!$A$8:$A${last_category_row}"),
        (f'$C$2:$C${ROWS + 1}', f"'{INFO}'!$E$12:$E${11 + len(cat.NATURES)}"),
        (f'$B$2:$B${ROWS + 1}', f"'{INFO}'!$E$8:$E$9"),
    ]
    for target, source in validations:
        dv = DataValidation(type='list', formula1=f'={source}', allow_blank=True)
        dv.error = '목록에 있는 값만 넣을 수 있습니다.'
        dv.errorTitle = '잘못된 값'
        ws.add_data_validation(dv)
        dv.add(target)

    return ws


def _sheet_monthly(wb, transactions):
    ws = wb.create_sheet(MONTHLY)
    months = aggregate.months_of(transactions) or ['2026-01']
    entries = [c for c in cat.ORDER if c != '미분류'] + ['미분류']

    _widths(ws, [16] + [14] * len(months))
    ws['A1'] = '월 × 카테고리 (수식 전용 — 손대지 마세요)'
    ws['A1'].font = Font(bold=True, size=12, color=NAVY)

    _header(ws, 3, ['카테고리'] + [f'{m[:4]}-{m[5:]}' for m in months])
    amount = f"'{TX}'!$F$2:$F${ROWS + 1}"
    category_col = f"'{TX}'!$D$2:$D${ROWS + 1}"
    monthkey = f"'{TX}'!$I$2:$I${ROWS + 1}"

    for i, category in enumerate(entries):
        row = 4 + i
        ws.cell(row=row, column=1, value=category).border = BORDER
        for j, month in enumerate(months):
            eom = f'DATE({month[:4]},{int(month[5:])},1)'
            cell = ws.cell(row=row, column=2 + j)
            cell.value = (f'=SUMIFS({amount},{category_col},$A{row},'
                          f'{monthkey},EOMONTH({eom},0))')
            cell.number_format = MONEY
            cell.border = BORDER

    return ws, months, entries


def _sheet_dashboard(wb, months, entries, settings):
    ws = wb.create_sheet(DASH)
    _widths(ws, [16, 16, 16, 16, 4, 18, 14, 14])
    ws.sheet_view.showGridLines = False

    ws['A1'] = '대시보드'
    ws['A1'].font = Font(bold=True, size=16, color=NAVY)
    ws['A2'] = f"기준월은 {INFO} 시트에서 바꿉니다."
    ws['A2'].font = Font(size=9, color='FF898781')

    amount = f"'{TX}'!$F$2:$F${ROWS + 1}"
    nature_col = f"'{TX}'!$C$2:$C${ROWS + 1}"
    category_col = f"'{TX}'!$D$2:$D${ROWS + 1}"
    monthkey = f"'{TX}'!$I$2:$I${ROWS + 1}"
    base = f"EOMONTH('{INFO}'!$B$3,0)"

    def by_nature(nature):
        return f'SUMIFS({amount},{nature_col},"{nature}",{monthkey},{base})'

    income = by_nature(cat.INCOME)
    fixed = by_nature(cat.FIXED)
    variable = by_nature(cat.VARIABLE)
    saving = by_nature(cat.SAVING)

    _header(ws, 4, ['수입', '지출', '저축·투자', '저축률'])
    values = [
        f'={income}',
        f'={fixed}+{variable}',
        f'={saving}',
        f'=IFERROR(({income}-({fixed}+{variable}))/{income},0)',
    ]
    for col, formula in enumerate(values, start=1):
        cell = ws.cell(row=5, column=col, value=formula)
        cell.number_format = PCT if col == 4 else MONEY
        cell.font = Font(bold=True, size=13)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    ws['A7'] = '고정비'
    ws['B7'] = f'={fixed}'
    ws['B7'].number_format = MONEY
    ws['A8'] = '변동비'
    ws['B8'] = f'={variable}'
    ws['B8'].number_format = MONEY
    ws['A9'] = '이번 달 잔액'
    ws['B9'] = f'={income}-({fixed}+{variable})-{saving}'
    ws['B9'].number_format = MONEY

    # 카테고리별 실적 / 예산 / 달성률
    _header(ws, 11, ['카테고리', '이번 달', '예산', '예산 대비'])
    last_info_row = 7 + len([c for c in cat.ORDER if c != '미분류'])
    for i, category in enumerate(entries):
        row = 12 + i
        ws.cell(row=row, column=1, value=category).border = BORDER

        actual = ws.cell(row=row, column=2)
        actual.value = (f'=SUMIFS({amount},{category_col},$A{row},'
                        f'{monthkey},{base})')
        actual.number_format = MONEY
        actual.border = BORDER

        budget = ws.cell(row=row, column=3)
        budget.value = (f"=IFERROR(INDEX('{INFO}'!$C$8:$C${last_info_row},"
                        f"MATCH($A{row},'{INFO}'!$A$8:$A${last_info_row},0)),0)")
        budget.number_format = MONEY
        budget.border = BORDER

        rate = ws.cell(row=row, column=4)
        rate.value = f'=IFERROR($B{row}/$C{row},0)'
        rate.number_format = PCT
        rate.border = BORDER

    # 고정비 미납 체크
    fixed_categories = cat.categories_of(cat.FIXED)
    _header(ws, 11, ['고정비 항목', '이번 달', '상태'], fill=NAVY_LIGHT)
    ws.cell(row=11, column=6, value='고정비 항목')
    for i, category in enumerate(fixed_categories):
        row = 12 + i
        ws.cell(row=row, column=6, value=category).border = BORDER
        spent = ws.cell(row=row, column=7)
        spent.value = (f'=SUMIFS({amount},{category_col},$F{row},'
                       f'{monthkey},{base})')
        spent.number_format = MONEY
        spent.border = BORDER
        status = ws.cell(row=row, column=8)
        status.value = (f'=IF(COUNTIFS({category_col},$F{row},'
                        f'{monthkey},{base})=0,"미납","OK")')
        status.border = BORDER

    return ws


def _sheet_assets(wb, assets):
    ws = wb.create_sheet(ASSETS)
    _widths(ws, [16, 34, 18, 18, 12])

    ws['A1'] = '자산 · 부채'
    ws['A1'].font = Font(bold=True, size=14, color=NAVY)

    _header(ws, 3, ['소유자', '항목', '총자산', '총부채', '순자산'])
    row = 4
    for owner, one in (assets or {}).items():
        ws.cell(row=row, column=1, value=owner).border = BORDER
        ws.cell(row=row, column=2, value='뱅크샐러드 재무현황').border = BORDER
        for col, key in ((3, 'assets'), (4, 'liabilities'), (5, 'net')):
            cell = ws.cell(row=row, column=col, value=int(one.get(key, 0)))
            cell.number_format = MONEY
            cell.border = BORDER
        row += 1

    if row > 4:
        ws.cell(row=row, column=2, value='합계').font = Font(bold=True)
        for col in (3, 4, 5):
            letter = get_column_letter(col)
            cell = ws.cell(row=row, column=col,
                           value=f'=SUM({letter}4:{letter}{row - 1})')
            cell.number_format = MONEY
            cell.font = Font(bold=True)
        row += 2

    investments = [i for one in (assets or {}).values() for i in one.get('investments', [])]
    if investments:
        _header(ws, row, ['종류', '상품명', '원금', '평가금액', '수익률'])
        start = row + 1
        for i, item in enumerate(investments):
            r = start + i
            ws.cell(row=r, column=1, value=item.get('group', '')).border = BORDER
            ws.cell(row=r, column=2, value=item.get('name', '')).border = BORDER
            ws.cell(row=r, column=3, value=int(item.get('원금', 0))).number_format = MONEY
            ws.cell(row=r, column=4, value=int(item.get('평가금액', 0))).number_format = MONEY
            rate = ws.cell(row=r, column=5, value=f'=IFERROR(($D{r}-$C{r})/$C{r},0)')
            rate.number_format = PCT
        row = start + len(investments) + 1

    loans = [l for one in (assets or {}).values() for l in one.get('loans', [])]
    if loans:
        _header(ws, row, ['종류', '상품명', '원금', '잔액'])
        start = row + 1
        for i, item in enumerate(loans):
            r = start + i
            ws.cell(row=r, column=1, value=item.get('group', '')).border = BORDER
            ws.cell(row=r, column=2, value=item.get('name', '')).border = BORDER
            ws.cell(row=r, column=3, value=int(item.get('원금', 0))).number_format = MONEY
            ws.cell(row=r, column=4, value=int(item.get('잔액', 0))).number_format = MONEY

    return ws


def build_workbook(transactions, assets, settings, path):
    """올인원 관리표를 만들어 path에 저장한다."""
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_usage(wb, settings)
    _sheet_info(wb, transactions, settings)
    _sheet_transactions(wb, transactions, settings)
    _, months, entries = _sheet_monthly(wb, transactions)
    _sheet_dashboard(wb, months, entries, settings)
    _sheet_assets(wb, assets)

    wb.save(path)
    return path
